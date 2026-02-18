import openpyxl as xl
from openpyxl.chart import BarChart,Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
             min_row=2,
             max_row=sheet.max_row,
             min_col=4,
             max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)









# from pathlib import Path
#
# path = Path()
# for file in path.glob('*'):
#     print(file)




# import converter
# from converter import lbs_to_kg
# lbs_to_kg(100)
# print(converter.kg_to_lbs(45))

# from utils import biggest_number

# num = [10, 3, 6, 2, 8, 4]
# biggest_number =  biggest_number(num)
# print(biggest_number)

# from ecommerce.shipping import calc_shipping
# calc_shipping()
# import random

# message = ["zane", "james", "dami", "emma"]
# word = random.choice(message)
# print(word)


# class Dice:
#     def roll(self):
#         first = random.randint(1, 8)
#         second = random.randint(1, 8)
#         return first,second


# dice = Dice()
# print(dice.roll())
# for i in range(3):
#     print(random.randint(10, 20))