from flask import Flask, render_template, request, jsonify
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

app = Flask(__name__)

# Logic from original EXCEL.py
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        if cell.value:
            corrected_price = cell.value * 0.9
            corrected_price_cell = sheet.cell(row, 4)
            corrected_price_cell.value = corrected_price

    values = Reference(sheet,
             min_row=2,
             max_row=sheet.max_row,
             min_col=4,
             max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
    return True

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process_excel', methods=['POST'])
def process_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'})
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Please upload an Excel file (.xlsx)'})
        
        filename = file.filename
        file.save(filename)
        
        process_workbook(filename)
        output_filename = f'processed_{filename}'
        
        import os
        if os.path.exists(output_filename):
            os.remove(output_filename)
        os.rename(filename, output_filename)
        
        return jsonify({'success': True, 'filename': output_filename})
    
    except Exception as e:
        return jsonify({'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True)
